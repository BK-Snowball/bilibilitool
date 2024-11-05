import requests
from bs4 import BeautifulSoup
import openpyxl as op
import re

# 识别码等
head = {
    "User-Agent": "修改你的识别码"}

response = requests.get(f"https://www.bilibili.com/video/修改你的BV号", headers=head)  # BV号
# 变量定义
select_fun = 0          # 方案选择参数
count3 = 1
count4 = 1
wb2 = op.Workbook()     # 创建工作簿对象
ws2 = wb2['Sheet']      # 创建子表
time_sec = 0
time_min = 0
time_sec_sum = 0
time_min_sum = 0
start_num = 1           # 起始集数
end_num = 48            # 结束集数
count_hour = 0
count_min = 0


# 本方案适用于分集格式 最后得到：分秒
def fun1():
    global response, count3, count4, ws2, time_sec, time_min
    global time_sec_sum, time_min_sum, start_num, end_num
    html = response.text
    soup = BeautifulSoup(html, "html.parser")
    # 分离所需信息
    all_titles = soup.findAll("div", attrs={"class": "video-episode-card__info-title"})  
    all_times = soup.findAll("div", attrs={"class": "video-episode-card__info-duration"})
    # 写入Excel
    for title_raw1 in all_titles:
        title_raw2 = title_raw1['title']                    
        ws2.cell(row=count3, column=1, value=title_raw2)    
        count3 += 1
    # 时长求和
    for time_raw1 in all_times:
        time_raw2 = time_raw1.string.strip()
        ws2.cell(row=count4, column=2, value=time_raw2)     
        # 选择时间计算
        if (count4 >= start_num) and (count4 <= end_num):
            if time_raw2[2] == ':':                       # 分秒    格式：11:22
                time_min = int(time_raw2[0:2])
                time_sec = int(time_raw2[3:5])
                time_sec_sum = time_sec_sum + time_sec
                time_min_sum = time_min_sum + time_min
            elif time_raw2[1] == ':':                     # 时分秒 格式：1:22:33
                time_min = int(time_raw2[2:4])
                time_sec = int(time_raw2[5:7])
                time_sec_sum = time_sec_sum + time_sec
                time_min_sum = time_min_sum + time_min + int(time_raw2[0:1]) * 60
        count4 += 1


# 本方案适用于分P格式 最后得到：秒
def fun2():
    global response, count3, count4, ws2, time_sec, time_min
    global time_sec_sum, time_min_sum, start_num, end_num
    html = response.text
    # 分离所需信息
    all_titles = re.findall(r'"part":"(.*?)"', html)            # 正则表达式提取所有标题
    all_times_temp1 = re.findall(r'"page".*?\}', html)          
    all_times_temp2 = "".join(all_times_temp1)                  
    all_times = re.findall(r'(?<=,)"duration":(\d+)(?=,)', all_times_temp2)  # 取出时间，单位秒
    # 写入Excel
    for title_raw1 in all_titles:
        title_raw2 = title_raw1                             # 从列表逐个取出
        ws2.cell(row=count3, column=1, value=title_raw2)    # 逐个写入
        count3 += 1
    for time_raw1 in all_times:
        time_raw2 = time_raw1                               
        ws2.cell(row=count4, column=2, value=time_raw2)     
        # 时长求和
        if (count4 >= start_num) and (count4 <= end_num):
            time_sec = int(time_raw2)  
            time_sec_sum = time_sec_sum + time_sec  # 计算所有视频秒和
        count4 += 1

# 方案选择
if select_fun != 0:
    if select_fun == 1:    # 分集格式
        fun1()
    elif select_fun == 2:  # 分P格式
        fun2()
elif select_fun == 0:
    print("您没有选择方案，将默认尝试分集方案（如果结果为0则尝试分P）")
    select_fun = 1
    fun1()
    if count_hour == 0 and count_min == 0 and time_sec_sum == 0:      # 如果分P无法得到结果则选择分集
        select_fun = 2
        fun2()
        if count_hour == 0 and count_min == 0 and time_sec_sum == 0:  # 如果分集也无法得到正确数据
            select_fun = 3

# 分秒和时分计算
while time_sec_sum >= 60:
    time_min_sum += 1
    time_sec_sum -= 60
    count_min = time_min_sum
else:
    count_min = time_min_sum 

while count_min >= 60:
    count_hour += 1
    count_min -= 60
# 结果输出
wb2.save('bilibilitool.xlsx')
if select_fun == 1:
    print("方案1：所选合集视频第{0}集到{1}集总时长为{2}分{3}秒。详见同目录下的xlsx文件".format(start_num, end_num, time_min_sum, time_sec_sum))
elif select_fun == 2:
    print("方案2：所选分P视频第{0}集到{1}集总时长为{2}分{3}秒。详见同目录下的xlsx文件".format(start_num, end_num, time_min_sum, time_sec_sum))
elif select_fun == 3:
    print("您输入的集数有误，请检查")

print("即{0}时{1}分{2}秒".format(count_hour, count_min, time_sec_sum))
